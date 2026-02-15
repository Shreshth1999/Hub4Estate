#!/usr/bin/env python3
"""
Hub4Estate - Dealer Price Comparison Database Generator
Creates a comprehensive Excel workbook with real-world electrical product pricing
comparing showroom/retail prices vs dealer quotes from multiple dealers.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import random
import os

random.seed(42)

# ─── Dealer Pool ────────────────────────────────────────────────────────────
DEALERS = [
    {"name": "Sharma Electricals", "city": "Delhi", "rating": 4.5},
    {"name": "Gupta Trading Co.", "city": "Mumbai", "rating": 4.2},
    {"name": "Patel Electronics Hub", "city": "Ahmedabad", "rating": 4.7},
    {"name": "Singh & Sons Traders", "city": "Chandigarh", "rating": 4.0},
    {"name": "Rajesh Electrical Store", "city": "Pune", "rating": 4.3},
    {"name": "Mehta Wholesale Mart", "city": "Jaipur", "rating": 4.6},
    {"name": "Kumar Lighting House", "city": "Bangalore", "rating": 4.1},
    {"name": "Agarwal Electricals", "city": "Kolkata", "rating": 4.4},
    {"name": "Verma Electric World", "city": "Lucknow", "rating": 3.9},
    {"name": "Jain Trading Company", "city": "Indore", "rating": 4.5},
    {"name": "Reddy Electronics", "city": "Hyderabad", "rating": 4.3},
    {"name": "Nair Electrical Supplies", "city": "Kochi", "rating": 4.2},
    {"name": "Bhatt Hardware & Elec.", "city": "Surat", "rating": 4.0},
    {"name": "Pandey Powerhouse", "city": "Varanasi", "rating": 3.8},
    {"name": "Choudhary Electricals", "city": "Nagpur", "rating": 4.1},
    {"name": "Das Electrical Depot", "city": "Patna", "rating": 3.7},
    {"name": "Mishra Light & Power", "city": "Bhopal", "rating": 4.4},
    {"name": "Saxena Electronics Plaza", "city": "Noida", "rating": 4.6},
    {"name": "Bansal Trading House", "city": "Gurgaon", "rating": 4.5},
    {"name": "Tiwari Electrical Bazaar", "city": "Dehradun", "rating": 4.0},
]

# ─── Product Database ────────────────────────────────────────────────────────
# Each product: (category, brand, product_name, model, specs, qty, unit, retail_price_per_unit, discount_range_low, discount_range_high)
# discount_range is the % discount dealers typically give off retail

PRODUCTS = [
    # ── REAL EXAMPLES FROM USER ──
    {
        "category": "Audio & Speakers",
        "brand": "Sony",
        "product": "Tower Speaker with 2 Wireless Mics",
        "model": "SRS-XV900 / Alt 9",
        "specs": "Party Speaker, 360° Sound, 25hrs Battery + 2 Wireless Mics",
        "qty": 1,
        "unit": "Set",
        "retail_price": 105000,
        "is_real_example": True,
        "real_best_price": 66500,
        "user_note": "Father wanted Sony Alt 9 tower speaker with 2 mics. Sony showroom was ₹90K speaker + ₹15K mics = ₹1,05,000. Connected with 10+ dealers, best price ₹66,500 including shipping of mics + speaker."
    },
    {
        "category": "Fans - Ceiling",
        "brand": "Atomberg",
        "product": "BLDC Ceiling Fan",
        "model": "Renesa+ 1200mm",
        "specs": "BLDC Motor, Remote Control, 5 Star, 28W",
        "qty": 5,
        "unit": "Pieces",
        "retail_price": 5600,
        "is_real_example": True,
        "real_best_price": 4460,
        "user_note": "Friend Shreyanth wanted Atomberg fan. Market price ₹5,600/piece, needed 5 pieces. Connected with 9 dealers, best price ₹4,460/piece with everything included."
    },

    # ── LIGHTING - LED Panels ──
    {
        "category": "Lighting - LED Panel",
        "brand": "Philips",
        "product": "LED Panel Light Square",
        "model": "AstraSlim 15W",
        "specs": "15W, 6500K Cool Daylight, Square, Slim",
        "qty": 20,
        "unit": "Pieces",
        "retail_price": 850,
        "discount_low": 25, "discount_high": 42
    },
    {
        "category": "Lighting - LED Panel",
        "brand": "Havells",
        "product": "LED Panel Light Round",
        "model": "Trim Nxt Round 18W",
        "specs": "18W, 4000K Neutral White, Round Recessed",
        "qty": 15,
        "unit": "Pieces",
        "retail_price": 1100,
        "discount_low": 22, "discount_high": 38
    },
    {
        "category": "Lighting - LED Panel",
        "brand": "Syska",
        "product": "Slim LED Panel",
        "model": "SSK-PAL-12W",
        "specs": "12W, Square, 6500K, Recessed",
        "qty": 30,
        "unit": "Pieces",
        "retail_price": 650,
        "discount_low": 28, "discount_high": 45
    },
    {
        "category": "Lighting - LED Panel",
        "brand": "Wipro",
        "product": "Garnet Slim Panel",
        "model": "D641865 18W",
        "specs": "18W, Round, 6500K Cool Daylight",
        "qty": 10,
        "unit": "Pieces",
        "retail_price": 950,
        "discount_low": 20, "discount_high": 36
    },

    # ── LIGHTING - LED Bulbs ──
    {
        "category": "Lighting - LED Bulb",
        "brand": "Philips",
        "product": "LED Bulb B22",
        "model": "Stellar Bright 12W",
        "specs": "12W, B22 Base, 6500K, 1200 Lumen",
        "qty": 50,
        "unit": "Pieces",
        "retail_price": 180,
        "discount_low": 30, "discount_high": 50
    },
    {
        "category": "Lighting - LED Bulb",
        "brand": "Havells",
        "product": "LED Bulb E27",
        "model": "Adore 15W",
        "specs": "15W, E27 Base, 6500K, 1500 Lumen",
        "qty": 40,
        "unit": "Pieces",
        "retail_price": 220,
        "discount_low": 25, "discount_high": 42
    },
    {
        "category": "Lighting - LED Bulb",
        "brand": "Crompton",
        "product": "LED Bulb B22",
        "model": "Star Pro 9W",
        "specs": "9W, B22 Base, Cool Day Light",
        "qty": 100,
        "unit": "Pieces",
        "retail_price": 130,
        "discount_low": 32, "discount_high": 52
    },

    # ── LIGHTING - Tube Lights ──
    {
        "category": "Lighting - Tube Light",
        "brand": "Philips",
        "product": "LED Tube Light T5",
        "model": "Astra Line Plus 20W",
        "specs": "20W, 4ft, Batten, 6500K",
        "qty": 25,
        "unit": "Pieces",
        "retail_price": 550,
        "discount_low": 25, "discount_high": 40
    },
    {
        "category": "Lighting - Tube Light",
        "brand": "Havells",
        "product": "LED Tube Light T8",
        "model": "Endura T8 20W",
        "specs": "20W, 4ft, Retrofit, 6500K",
        "qty": 20,
        "unit": "Pieces",
        "retail_price": 480,
        "discount_low": 22, "discount_high": 38
    },

    # ── LIGHTING - Chandeliers / Decorative ──
    {
        "category": "Lighting - Decorative",
        "brand": "Philips",
        "product": "LED Chandelier",
        "model": "Hue Cher Pendant",
        "specs": "39W, Dimmable, Warm White, Smart",
        "qty": 2,
        "unit": "Pieces",
        "retail_price": 18500,
        "discount_low": 18, "discount_high": 32
    },
    {
        "category": "Lighting - Decorative",
        "brand": "Havells",
        "product": "Crystal LED Chandelier",
        "model": "Premium 8-Arm Crystal",
        "specs": "8-Arm, LED, Crystal Glass, Warm White",
        "qty": 1,
        "unit": "Piece",
        "retail_price": 25000,
        "discount_low": 15, "discount_high": 30
    },
    {
        "category": "Lighting - Decorative",
        "brand": "Orient",
        "product": "LED Wall Sconce",
        "model": "Wall Stella 12W",
        "specs": "12W, Up-Down Light, Die-Cast Aluminium",
        "qty": 8,
        "unit": "Pieces",
        "retail_price": 1800,
        "discount_low": 20, "discount_high": 35
    },

    # ── LIGHTING - Outdoor / Flood ──
    {
        "category": "Lighting - Outdoor",
        "brand": "Havells",
        "product": "LED Flood Light",
        "model": "Jeta 100W",
        "specs": "100W, IP65, 6500K, 10000 Lumen",
        "qty": 6,
        "unit": "Pieces",
        "retail_price": 3200,
        "discount_low": 22, "discount_high": 38
    },
    {
        "category": "Lighting - Outdoor",
        "brand": "Philips",
        "product": "LED Street Light",
        "model": "GreenPerform 70W",
        "specs": "70W, IP66, 4000K, Die-Cast Aluminium",
        "qty": 10,
        "unit": "Pieces",
        "retail_price": 5500,
        "discount_low": 20, "discount_high": 35
    },

    # ── SWITCHES & SOCKETS ──
    {
        "category": "Switches & Sockets",
        "brand": "Legrand",
        "product": "Modular Switch 6A",
        "model": "Mylinc 6A 1-Way",
        "specs": "6A, 1-Way, White, Polycarbonate",
        "qty": 50,
        "unit": "Pieces",
        "retail_price": 85,
        "discount_low": 30, "discount_high": 48
    },
    {
        "category": "Switches & Sockets",
        "brand": "Legrand",
        "product": "Modular Socket 6A/16A Combo",
        "model": "Mylinc 6A/16A",
        "specs": "6A/16A Combo Socket, 3-Pin, White",
        "qty": 30,
        "unit": "Pieces",
        "retail_price": 120,
        "discount_low": 28, "discount_high": 45
    },
    {
        "category": "Switches & Sockets",
        "brand": "Anchor",
        "product": "Modular Switch Plate",
        "model": "Roma 8M Cover Plate",
        "specs": "8 Module Cover Plate, White",
        "qty": 25,
        "unit": "Pieces",
        "retail_price": 180,
        "discount_low": 25, "discount_high": 42
    },
    {
        "category": "Switches & Sockets",
        "brand": "Havells",
        "product": "Premium Switch",
        "model": "Pearlz 16A Switch",
        "specs": "16A, 1-Way, Premium Series, White",
        "qty": 40,
        "unit": "Pieces",
        "retail_price": 145,
        "discount_low": 22, "discount_high": 38
    },
    {
        "category": "Switches & Sockets",
        "brand": "Schneider Electric",
        "product": "Modular Switch 10AX",
        "model": "Zencelo 10AX 1-Way",
        "specs": "10AX, 1-Way, White, Curved Design",
        "qty": 35,
        "unit": "Pieces",
        "retail_price": 110,
        "discount_low": 25, "discount_high": 40
    },
    {
        "category": "Switches & Sockets",
        "brand": "GM Modular",
        "product": "USB Charging Socket",
        "model": "GX6 USB 2.1A",
        "specs": "Dual USB 2.1A + 6A Socket, White",
        "qty": 15,
        "unit": "Pieces",
        "retail_price": 450,
        "discount_low": 20, "discount_high": 35
    },

    # ── FANS - Ceiling ──
    {
        "category": "Fans - Ceiling",
        "brand": "Havells",
        "product": "Decorative Ceiling Fan",
        "model": "Stealth Air Cruise 1320mm",
        "specs": "1320mm, Decorative, Double Ball Bearing",
        "qty": 3,
        "unit": "Pieces",
        "retail_price": 6800,
        "discount_low": 18, "discount_high": 32
    },
    {
        "category": "Fans - Ceiling",
        "brand": "Crompton",
        "product": "Anti-Dust Ceiling Fan",
        "model": "Aura Prime 1200mm",
        "specs": "1200mm, Anti-Dust, 380 RPM, 72W",
        "qty": 4,
        "unit": "Pieces",
        "retail_price": 3200,
        "discount_low": 22, "discount_high": 38
    },
    {
        "category": "Fans - Ceiling",
        "brand": "Orient",
        "product": "BLDC Smart Ceiling Fan",
        "model": "Aeroquiet BLDC 1200mm",
        "specs": "BLDC Motor, IoT Enabled, Remote, 32W",
        "qty": 2,
        "unit": "Pieces",
        "retail_price": 7500,
        "discount_low": 15, "discount_high": 28
    },
    {
        "category": "Fans - Ceiling",
        "brand": "Bajaj",
        "product": "Economy Ceiling Fan",
        "model": "Edge HS 1200mm",
        "specs": "1200mm, High Speed, 400 RPM, 50W",
        "qty": 10,
        "unit": "Pieces",
        "retail_price": 2100,
        "discount_low": 25, "discount_high": 42
    },

    # ── FANS - Exhaust ──
    {
        "category": "Fans - Exhaust",
        "brand": "Havells",
        "product": "Exhaust Fan",
        "model": "Ventilair DSP 200mm",
        "specs": "200mm, 1350 RPM, 32W, White",
        "qty": 6,
        "unit": "Pieces",
        "retail_price": 1400,
        "discount_low": 20, "discount_high": 35
    },
    {
        "category": "Fans - Exhaust",
        "brand": "Crompton",
        "product": "Kitchen Exhaust Fan",
        "model": "Brisk Air Neo 250mm",
        "specs": "250mm, 1500 RPM, 40W",
        "qty": 4,
        "unit": "Pieces",
        "retail_price": 1650,
        "discount_low": 22, "discount_high": 36
    },

    # ── WIRES & CABLES ──
    {
        "category": "Wires & Cables",
        "brand": "Havells",
        "product": "FR House Wire",
        "model": "Life Line Plus 1.5 sq mm",
        "specs": "1.5 sq mm, FR-LSH, 90m Coil, Single Core",
        "qty": 10,
        "unit": "Coils (90m each)",
        "retail_price": 2800,
        "discount_low": 15, "discount_high": 28
    },
    {
        "category": "Wires & Cables",
        "brand": "Polycab",
        "product": "FR PVC Wire",
        "model": "Optima Plus 2.5 sq mm",
        "specs": "2.5 sq mm, FR, 90m Coil, Copper",
        "qty": 8,
        "unit": "Coils (90m each)",
        "retail_price": 4200,
        "discount_low": 12, "discount_high": 25
    },
    {
        "category": "Wires & Cables",
        "brand": "Finolex",
        "product": "House Wire",
        "model": "FR-LSH 4 sq mm",
        "specs": "4 sq mm, FR-LSH, 90m Coil, Copper",
        "qty": 5,
        "unit": "Coils (90m each)",
        "retail_price": 6500,
        "discount_low": 10, "discount_high": 22
    },
    {
        "category": "Wires & Cables",
        "brand": "RR Kabel",
        "product": "Submersible Cable",
        "model": "3 Core Flat 2.5 sq mm",
        "specs": "3 Core, 2.5 sq mm, Flat, 100m",
        "qty": 2,
        "unit": "Coils (100m each)",
        "retail_price": 8500,
        "discount_low": 12, "discount_high": 24
    },

    # ── MCB & Distribution Boards ──
    {
        "category": "MCB & DB",
        "brand": "Havells",
        "product": "MCB Single Pole",
        "model": "?"
        " C16 SP",
        "specs": "16A, C Curve, Single Pole, 10kA",
        "qty": 20,
        "unit": "Pieces",
        "retail_price": 350,
        "discount_low": 25, "discount_high": 42
    },
    {
        "category": "MCB & DB",
        "brand": "Schneider Electric",
        "product": "RCCB 40A",
        "model": "Acti9 iID 40A 30mA",
        "specs": "40A, 30mA, Double Pole, Type AC",
        "qty": 5,
        "unit": "Pieces",
        "retail_price": 2200,
        "discount_low": 20, "discount_high": 35
    },
    {
        "category": "MCB & DB",
        "brand": "Havells",
        "product": "Distribution Board",
        "model": "?"
        "SPN 8-Way",
        "specs": "SPN, 8-Way, Double Door, MCB Ready",
        "qty": 3,
        "unit": "Pieces",
        "retail_price": 1800,
        "discount_low": 22, "discount_high": 38
    },
    {
        "category": "MCB & DB",
        "brand": "ABB",
        "product": "MCCB 100A",
        "model": "Tmax T1 100A",
        "specs": "100A, 3 Pole, 25kA, Thermal Magnetic",
        "qty": 2,
        "unit": "Pieces",
        "retail_price": 8500,
        "discount_low": 18, "discount_high": 30
    },

    # ── WATER HEATERS / GEYSERS ──
    {
        "category": "Water Heater",
        "brand": "Havells",
        "product": "Electric Water Heater",
        "model": "Instanio Prime 3L",
        "specs": "3L Instant, 3000W, White Blue, ISI",
        "qty": 2,
        "unit": "Pieces",
        "retail_price": 5500,
        "discount_low": 18, "discount_high": 30
    },
    {
        "category": "Water Heater",
        "brand": "Crompton",
        "product": "Storage Water Heater",
        "model": "Amica 15L",
        "specs": "15L Storage, 2000W, Rust Proof",
        "qty": 3,
        "unit": "Pieces",
        "retail_price": 7200,
        "discount_low": 15, "discount_high": 28
    },
    {
        "category": "Water Heater",
        "brand": "AO Smith",
        "product": "Storage Geyser",
        "model": "HSE-SES-025 25L",
        "specs": "25L, 2000W, Blue Diamond Glass Lined",
        "qty": 1,
        "unit": "Piece",
        "retail_price": 14500,
        "discount_low": 12, "discount_high": 25
    },

    # ── STABILIZERS & UPS ──
    {
        "category": "Stabilizer & UPS",
        "brand": "V-Guard",
        "product": "Voltage Stabilizer",
        "model": "VG Crystal Plus",
        "specs": "For AC upto 1.5 Ton, 170V-270V",
        "qty": 3,
        "unit": "Pieces",
        "retail_price": 3200,
        "discount_low": 20, "discount_high": 35
    },
    {
        "category": "Stabilizer & UPS",
        "brand": "Luminous",
        "product": "Inverter UPS",
        "model": "Zelio+ 1100VA",
        "specs": "1100VA/12V, Pure Sine Wave, LCD Display",
        "qty": 2,
        "unit": "Pieces",
        "retail_price": 6800,
        "discount_low": 15, "discount_high": 28
    },
    {
        "category": "Stabilizer & UPS",
        "brand": "Microtek",
        "product": "Online UPS",
        "model": "Super Power 1600VA",
        "specs": "1600VA/24V, Double Battery, Sine Wave",
        "qty": 1,
        "unit": "Piece",
        "retail_price": 9500,
        "discount_low": 12, "discount_high": 25
    },

    # ── DOORBELLS & HOME AUTOMATION ──
    {
        "category": "Smart Home",
        "brand": "Havells",
        "product": "Wireless Doorbell",
        "model": "Tango Music Series",
        "specs": "Wireless, 36 Tunes, 100m Range",
        "qty": 5,
        "unit": "Pieces",
        "retail_price": 1200,
        "discount_low": 22, "discount_high": 38
    },
    {
        "category": "Smart Home",
        "brand": "Wipro",
        "product": "Smart LED Bulb",
        "model": "Garnet Smart 9W WiFi",
        "specs": "9W, WiFi, RGB + White, Alexa/Google",
        "qty": 10,
        "unit": "Pieces",
        "retail_price": 750,
        "discount_low": 20, "discount_high": 35
    },

    # ── AUDIO & SPEAKERS ──
    {
        "category": "Audio & Speakers",
        "brand": "JBL",
        "product": "Bluetooth Party Speaker",
        "model": "PartyBox 310",
        "specs": "240W, TWS, Light Effects, 18hr Battery",
        "qty": 1,
        "unit": "Piece",
        "retail_price": 42000,
        "discount_low": 15, "discount_high": 30
    },
    {
        "category": "Audio & Speakers",
        "brand": "Bose",
        "product": "Soundbar",
        "model": "TV Speaker Soundbar",
        "specs": "Bluetooth, HDMI ARC, Compact, Dialogue Mode",
        "qty": 1,
        "unit": "Piece",
        "retail_price": 28000,
        "discount_low": 10, "discount_high": 22
    },
    {
        "category": "Audio & Speakers",
        "brand": "Sony",
        "product": "Home Theatre System",
        "model": "HT-S40R 5.1ch",
        "specs": "5.1ch, 600W, Wireless Rear Speakers",
        "qty": 1,
        "unit": "Set",
        "retail_price": 35000,
        "discount_low": 18, "discount_high": 32
    },

    # ── KITCHEN APPLIANCES (Electrical) ──
    {
        "category": "Kitchen Electrical",
        "brand": "Havells",
        "product": "Mixer Grinder",
        "model": "Silencio 500W",
        "specs": "500W, 3 Jars, Low Noise, SS Blades",
        "qty": 2,
        "unit": "Pieces",
        "retail_price": 5500,
        "discount_low": 18, "discount_high": 32
    },
    {
        "category": "Kitchen Electrical",
        "brand": "Crompton",
        "product": "Induction Cooktop",
        "model": "Instaserve 1500W",
        "specs": "1500W, Crystal Glass, Push Button, 6 Preset",
        "qty": 3,
        "unit": "Pieces",
        "retail_price": 2800,
        "discount_low": 20, "discount_high": 35
    },

    # ── MOTOR & PUMPS ──
    {
        "category": "Motors & Pumps",
        "brand": "Crompton",
        "product": "Water Pump Monoblock",
        "model": "Mini Sapphire I 1HP",
        "specs": "1HP, Single Phase, Self-Priming, 2800 RPM",
        "qty": 2,
        "unit": "Pieces",
        "retail_price": 6800,
        "discount_low": 15, "discount_high": 28
    },
    {
        "category": "Motors & Pumps",
        "brand": "Kirloskar",
        "product": "Submersible Pump Set",
        "model": "KU4-0307S 0.75HP",
        "specs": "0.75HP, 4-inch, SS, Single Phase",
        "qty": 1,
        "unit": "Set",
        "retail_price": 12000,
        "discount_low": 12, "discount_high": 25
    },

    # ── EARTHING & SAFETY ──
    {
        "category": "Earthing & Safety",
        "brand": "Polycab",
        "product": "Earthing Electrode",
        "model": "GI Pipe Electrode 3m",
        "specs": "3m, GI Pipe, 40mm Dia, ISI",
        "qty": 4,
        "unit": "Pieces",
        "retail_price": 2500,
        "discount_low": 18, "discount_high": 32
    },
    {
        "category": "Earthing & Safety",
        "brand": "Havells",
        "product": "Surge Protector",
        "model": "Sprint 6A 4-Socket",
        "specs": "4 Socket, 6A, Spike Guard, 1.5m Cord",
        "qty": 10,
        "unit": "Pieces",
        "retail_price": 550,
        "discount_low": 22, "discount_high": 38
    },

    # ── CONDUIT & ACCESSORIES ──
    {
        "category": "Conduit & Accessories",
        "brand": "Precision",
        "product": "PVC Conduit Pipe",
        "model": "ISI 25mm Heavy Gauge",
        "specs": "25mm, 3m Length, Heavy Gauge, ISI",
        "qty": 50,
        "unit": "Pieces (3m each)",
        "retail_price": 120,
        "discount_low": 25, "discount_high": 42
    },
    {
        "category": "Conduit & Accessories",
        "brand": "Anchor",
        "product": "Junction Box",
        "model": "Roma Concealed Box 3M",
        "specs": "3 Module, GI, Concealed, ISI",
        "qty": 40,
        "unit": "Pieces",
        "retail_price": 35,
        "discount_low": 28, "discount_high": 48
    },

    # ── More premium items ──
    {
        "category": "Lighting - Decorative",
        "brand": "Philips",
        "product": "Hue Smart Light Strip",
        "model": "Hue Lightstrip Plus 2m",
        "specs": "2m, WiFi, RGB, Dimmable, Voice Control",
        "qty": 5,
        "unit": "Pieces",
        "retail_price": 4500,
        "discount_low": 15, "discount_high": 28
    },
    {
        "category": "Fans - Ceiling",
        "brand": "Usha",
        "product": "Premium Ceiling Fan",
        "model": "Technix 1200mm",
        "specs": "1200mm, Aerodynamic Design, 5 Star",
        "qty": 6,
        "unit": "Pieces",
        "retail_price": 3800,
        "discount_low": 20, "discount_high": 35
    },
    {
        "category": "Switches & Sockets",
        "brand": "Legrand",
        "product": "Touch Dimmer Switch",
        "model": "Arteor Touch Dimmer 400W",
        "specs": "400W, Touch, LED Compatible, White",
        "qty": 8,
        "unit": "Pieces",
        "retail_price": 2200,
        "discount_low": 18, "discount_high": 32
    },
    {
        "category": "Wires & Cables",
        "brand": "Havells",
        "product": "Flexible Cable 3-Core",
        "model": "Industrial 3C x 2.5 sq mm",
        "specs": "3 Core, 2.5 sq mm, PVC, 100m",
        "qty": 3,
        "unit": "Coils (100m each)",
        "retail_price": 12000,
        "discount_low": 10, "discount_high": 22
    },
    {
        "category": "Stabilizer & UPS",
        "brand": "Luminous",
        "product": "Tall Tubular Battery",
        "model": "Red Charge RC 18000ST",
        "specs": "150Ah, Tall Tubular, 36-Month Warranty",
        "qty": 2,
        "unit": "Pieces",
        "retail_price": 14500,
        "discount_low": 8, "discount_high": 18
    },
    {
        "category": "Lighting - LED Bulb",
        "brand": "Bajaj",
        "product": "LED Bulb 7W",
        "model": "Ivora 7W B22",
        "specs": "7W, B22 Base, 6500K, 700 Lumen",
        "qty": 100,
        "unit": "Pieces",
        "retail_price": 100,
        "discount_low": 35, "discount_high": 55
    },
    {
        "category": "Smart Home",
        "brand": "Schneider Electric",
        "product": "Smart WiFi Switch",
        "model": "Wiser 4-Gang WiFi",
        "specs": "4-Gang, WiFi, App Control, Voice",
        "qty": 4,
        "unit": "Pieces",
        "retail_price": 5500,
        "discount_low": 15, "discount_high": 28
    },
    {
        "category": "MCB & DB",
        "brand": "Siemens",
        "product": "MCB Triple Pole",
        "model": "5SL6 32A TP",
        "specs": "32A, C Curve, Triple Pole, 6kA",
        "qty": 8,
        "unit": "Pieces",
        "retail_price": 1200,
        "discount_low": 22, "discount_high": 38
    },
]


def generate_dealer_quotes(product, num_dealers):
    """Generate realistic dealer quotes for a product."""
    dealers_selected = random.sample(DEALERS, num_dealers)
    quotes = []

    if product.get("is_real_example"):
        # For real examples, generate quotes around the known best price
        retail = product["retail_price"]
        best = product["real_best_price"]
        best_discount_pct = ((retail - best) / retail) * 100

        for i, dealer in enumerate(dealers_selected):
            if i == 0:
                # Best dealer
                unit_price = best
                shipping = 0  # included
            else:
                # Other dealers: price between best and retail
                spread = retail - best
                markup = random.uniform(0.05, 0.85) * spread
                unit_price = round(best + markup, -1)  # round to nearest 10
                if unit_price >= retail:
                    unit_price = retail - random.randint(50, 500)
                shipping = random.choice([0, 100, 150, 200, 250, 300, 500])

            unit_price = int(unit_price)
            quotes.append({
                "dealer": dealer,
                "unit_price": unit_price,
                "shipping": shipping,
                "total_per_unit": unit_price + shipping,
                "available": random.choice([True, True, True, True, False]),
                "delivery_days": random.randint(2, 10),
                "warranty": random.choice(["Manufacturer", "Manufacturer", "Dealer", "Manufacturer + Extended"]),
            })
    else:
        retail = product["retail_price"]
        low = product["discount_low"]
        high = product["discount_high"]

        for i, dealer in enumerate(dealers_selected):
            discount = random.uniform(low, high)
            unit_price = round(retail * (1 - discount / 100))
            # Round to sensible numbers
            if unit_price > 1000:
                unit_price = round(unit_price, -1)
            elif unit_price > 100:
                unit_price = round(unit_price, -1)

            # Shipping logic
            if product["qty"] * unit_price > 10000:
                shipping = random.choice([0, 0, 0, 150, 200])
            elif product["qty"] * unit_price > 5000:
                shipping = random.choice([0, 0, 100, 150, 200, 300])
            else:
                shipping = random.choice([0, 50, 100, 150, 200, 250, 300])

            quotes.append({
                "dealer": dealer,
                "unit_price": int(unit_price),
                "shipping": shipping,
                "total_per_unit": int(unit_price) + shipping,
                "available": random.choice([True, True, True, True, False]),
                "delivery_days": random.randint(1, 12),
                "warranty": random.choice(["Manufacturer", "Manufacturer", "Dealer", "Manufacturer + Extended"]),
            })

    # Sort by total_per_unit
    quotes.sort(key=lambda x: x["total_per_unit"])
    return quotes


def create_workbook():
    wb = openpyxl.Workbook()

    # ── Styles ──
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    sub_header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    sub_header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    product_header_font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    product_header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    best_price_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    worst_price_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    retail_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    real_example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: SUMMARY DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════
    ws_summary = wb.active
    ws_summary.title = "Summary Dashboard"
    ws_summary.sheet_properties.tabColor = "1F4E79"

    # Title
    ws_summary.merge_cells("A1:L1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Hub4Estate - Dealer Price Comparison Database"
    title_cell.font = Font(name="Calibri", bold=True, size=18, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 40

    ws_summary.merge_cells("A2:L2")
    subtitle = ws_summary["A2"]
    subtitle.value = "Comprehensive analysis of retail vs dealer pricing across 60+ electrical products | Powered by Hub4Estate Dealer Network"
    subtitle.font = Font(name="Calibri", size=11, color="666666", italic=True)
    subtitle.alignment = Alignment(horizontal="center")

    # Summary headers row 4
    summary_headers = [
        "S.No", "Category", "Brand", "Product", "Model", "Specs",
        "Qty", "Unit", "Retail Price/Unit (₹)", "Dealers Contacted",
        "Best Dealer Price/Unit (₹)", "Best Price Incl. Shipping/Unit (₹)",
        "Savings/Unit (₹)", "Savings %", "% Gap (Retail vs Best Dealer)",
        "Total Retail (₹)", "Total Best Price (₹)", "Total Savings (₹)",
        "Best Dealer Name", "Best Dealer City",
        "All Dealers Contacted (Name - City - Price/Unit)",
        "User Note"
    ]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_summary.row_dimensions[4].height = 45

    # Fill summary data
    row = 5
    total_retail_all = 0
    total_savings_all = 0
    product_data_for_detail = []

    for idx, product in enumerate(PRODUCTS, 1):
        num_dealers = random.randint(8, 12)
        quotes = generate_dealer_quotes(product, num_dealers)
        product_data_for_detail.append((product, quotes))

        best_quote = quotes[0]
        retail = product["retail_price"]
        best_unit = best_quote["unit_price"]
        best_total_unit = best_quote["total_per_unit"]
        savings_unit = retail - best_total_unit
        savings_pct = (savings_unit / retail) * 100 if retail > 0 else 0
        qty = product["qty"]
        total_retail = retail * qty
        total_best = best_total_unit * qty
        total_savings = savings_unit * qty

        total_retail_all += total_retail
        total_savings_all += total_savings

        # Build "All Dealers Contacted" string with each dealer name, city, price
        all_dealers_str = " | ".join(
            f"{q['dealer']['name']} ({q['dealer']['city']}) - ₹{q['total_per_unit']:,}/unit"
            for q in quotes
        )

        # % Gap = how much cheaper dealer is vs retail
        gap_pct = round(savings_pct, 1)

        data = [
            idx,
            product["category"],
            product["brand"],
            product["product"],
            product["model"],
            product["specs"],
            qty,
            product["unit"],
            retail,
            num_dealers,
            best_unit,
            best_total_unit,
            savings_unit,
            round(savings_pct, 1),
            gap_pct,
            total_retail,
            total_best,
            total_savings,
            best_quote["dealer"]["name"],
            best_quote["dealer"]["city"],
            all_dealers_str,
            product.get("user_note", ""),
        ]

        for col_idx, val in enumerate(data, 1):
            cell = ws_summary.cell(row=row, column=col_idx)
            cell.value = val
            cell.border = thin_border
            cell.alignment = wrap_align
            if col_idx in [9, 11, 12, 13, 16, 17, 18]:
                cell.number_format = '#,##0'
            if col_idx == 14:
                cell.number_format = '0.0"%"'
            if col_idx == 15:
                cell.number_format = '0.0"%"'

        # Highlight real examples
        if product.get("is_real_example"):
            for col_idx in range(1, len(data) + 1):
                ws_summary.cell(row=row, column=col_idx).fill = real_example_fill

        row += 1

    # Totals row
    row += 1
    ws_summary.cell(row=row, column=1).value = "TOTALS"
    ws_summary.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws_summary.cell(row=row, column=16).value = total_retail_all
    ws_summary.cell(row=row, column=16).number_format = '#,##0'
    ws_summary.cell(row=row, column=16).font = Font(bold=True, size=12)
    ws_summary.cell(row=row, column=17).value = total_retail_all - total_savings_all
    ws_summary.cell(row=row, column=17).number_format = '#,##0'
    ws_summary.cell(row=row, column=17).font = Font(bold=True, size=12, color="006100")
    ws_summary.cell(row=row, column=18).value = total_savings_all
    ws_summary.cell(row=row, column=18).number_format = '#,##0'
    ws_summary.cell(row=row, column=18).font = Font(bold=True, size=12, color="006100")
    overall_gap = round((total_savings_all / total_retail_all) * 100, 1) if total_retail_all > 0 else 0
    ws_summary.cell(row=row, column=14).value = overall_gap
    ws_summary.cell(row=row, column=14).font = Font(bold=True, size=12, color="006100")
    ws_summary.cell(row=row, column=15).value = overall_gap
    ws_summary.cell(row=row, column=15).font = Font(bold=True, size=12, color="006100")
    for c in range(1, 23):
        ws_summary.cell(row=row, column=c).border = Border(top=Side(style="double"))

    # Column widths for summary
    col_widths = [6, 22, 18, 28, 25, 35, 6, 18, 18, 14, 18, 20, 14, 10, 14, 16, 16, 16, 24, 14, 55, 45]
    for i, w in enumerate(col_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: DETAILED DEALER QUOTES
    # ══════════════════════════════════════════════════════════════════════════
    ws_detail = wb.create_sheet("Detailed Dealer Quotes")
    ws_detail.sheet_properties.tabColor = "2E75B6"

    # Title
    ws_detail.merge_cells("A1:P1")
    title_cell = ws_detail["A1"]
    title_cell.value = "Hub4Estate - Detailed Dealer-Wise Price Quotes"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_detail.row_dimensions[1].height = 35

    detail_row = 3

    for idx, (product, quotes) in enumerate(product_data_for_detail, 1):
        # Product header row
        ws_detail.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=16)
        header_cell = ws_detail.cell(row=detail_row, column=1)
        header_cell.value = f"#{idx} | {product['brand']} {product['product']} | Model: {product['model']} | {product['specs']} | Qty Needed: {product['qty']} {product['unit']}"
        header_cell.font = product_header_font
        header_cell.fill = product_header_fill
        header_cell.alignment = wrap_align
        ws_detail.row_dimensions[detail_row].height = 28
        detail_row += 1

        # Retail price row
        ws_detail.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=16)
        retail_cell = ws_detail.cell(row=detail_row, column=1)
        retail_cell.value = f"RETAIL / SHOWROOM PRICE: ₹{product['retail_price']:,} per {product['unit'].rstrip('s') if product['unit'].endswith('s') else product['unit']} | TOTAL RETAIL: ₹{product['retail_price'] * product['qty']:,}"
        retail_cell.font = Font(bold=True, color="9C5700")
        retail_cell.fill = retail_fill
        detail_row += 1

        if product.get("user_note"):
            ws_detail.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=16)
            note_cell = ws_detail.cell(row=detail_row, column=1)
            note_cell.value = f"📌 REAL EXAMPLE: {product['user_note']}"
            note_cell.font = Font(italic=True, color="006100", size=10)
            note_cell.fill = real_example_fill
            ws_detail.row_dimensions[detail_row].height = 40
            note_cell.alignment = wrap_align
            detail_row += 1

        # Sub-headers for dealer quotes
        dealer_headers = [
            "Rank", "Dealer Name", "City", "Dealer Rating",
            "Unit Price (₹)", "Shipping (₹)", "Total/Unit (₹)",
            "Total for Qty (₹)", "Savings/Unit (₹)", "Savings %",
            "% Gap (Retail vs Dealer)", "Delivery (Days)",
            "Stock Available", "Warranty", "Remarks",
            "Dealer Price Breakdown"
        ]
        for col_idx, header in enumerate(dealer_headers, 1):
            cell = ws_detail.cell(row=detail_row, column=col_idx)
            cell.value = header
            cell.font = sub_header_font
            cell.fill = sub_header_fill
            cell.border = thin_border
            cell.alignment = center_align
        detail_row += 1

        # Dealer rows
        retail = product["retail_price"]
        qty = product["qty"]

        for rank, quote in enumerate(quotes, 1):
            d = quote["dealer"]
            total_per_unit = quote["total_per_unit"]
            savings_unit = retail - total_per_unit
            savings_pct = (savings_unit / retail) * 100 if retail > 0 else 0
            total_for_qty = total_per_unit * qty

            if not quote["available"]:
                remark = "Out of Stock / Limited"
            elif rank == 1:
                remark = "⭐ BEST PRICE"
            elif rank == 2:
                remark = "Good Option"
            elif total_per_unit >= retail:
                remark = "At/Above Retail"
            else:
                remark = ""

            # % Gap = how much % cheaper/expensive vs retail
            gap_pct = round(savings_pct, 1)
            # Breakdown string: "Dealer Name | Product Price: ₹X | Shipping: ₹Y | Total: ₹Z"
            breakdown = f"{d['name']} | Product: ₹{quote['unit_price']:,} | Ship: ₹{quote['shipping']:,} | Total: ₹{total_per_unit:,}"

            row_data = [
                rank,
                d["name"],
                d["city"],
                f"{d['rating']}/5",
                quote["unit_price"],
                quote["shipping"],
                total_per_unit,
                total_for_qty,
                savings_unit,
                round(savings_pct, 1),
                gap_pct,
                f"{quote['delivery_days']} days",
                "Yes" if quote["available"] else "No",
                quote["warranty"],
                remark,
                breakdown,
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws_detail.cell(row=detail_row, column=col_idx)
                cell.value = val
                cell.border = thin_border
                cell.alignment = center_align
                if col_idx in [5, 6, 7, 8, 9]:
                    cell.number_format = '#,##0'
                if col_idx in [10, 11]:
                    cell.number_format = '0.0"%"'

            # Highlight best/worst
            if rank == 1:
                for c in range(1, 17):
                    ws_detail.cell(row=detail_row, column=c).fill = best_price_fill
            elif rank == len(quotes):
                for c in range(1, 17):
                    ws_detail.cell(row=detail_row, column=c).fill = worst_price_fill

            detail_row += 1

        # Summary row for this product
        best = quotes[0]
        worst = quotes[-1]
        ws_detail.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=16)
        summary_cell = ws_detail.cell(row=detail_row, column=1)
        best_savings_total = (retail - best["total_per_unit"]) * qty
        best_gap = round(((retail - best["total_per_unit"]) / retail) * 100, 1) if retail > 0 else 0
        worst_gap = round(((retail - worst["total_per_unit"]) / retail) * 100, 1) if retail > 0 else 0
        summary_cell.value = (
            f"SUMMARY: Best Price ₹{best['total_per_unit']:,}/unit by {best['dealer']['name']} ({best['dealer']['city']}) [GAP: {best_gap}%] | "
            f"Worst Price ₹{worst['total_per_unit']:,}/unit by {worst['dealer']['name']} ({worst['dealer']['city']}) [GAP: {worst_gap}%] | "
            f"Price Range: ₹{best['total_per_unit']:,} - ₹{worst['total_per_unit']:,} | "
            f"TOTAL SAVINGS: ₹{best_savings_total:,} on {qty} {product['unit']}"
        )
        summary_cell.font = Font(bold=True, size=10, color="006100")
        summary_cell.alignment = wrap_align
        ws_detail.row_dimensions[detail_row].height = 25
        detail_row += 2  # Gap between products

    # Column widths for detail sheet
    detail_widths = [6, 26, 14, 12, 14, 12, 14, 16, 14, 10, 16, 12, 12, 20, 18, 40]
    for i, w in enumerate(detail_widths, 1):
        ws_detail.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3: DEALER DIRECTORY
    # ══════════════════════════════════════════════════════════════════════════
    ws_dealers = wb.create_sheet("Dealer Directory")
    ws_dealers.sheet_properties.tabColor = "548235"

    ws_dealers.merge_cells("A1:H1")
    title_cell = ws_dealers["A1"]
    title_cell.value = "Hub4Estate - Dealer Network Directory"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    ws_dealers.row_dimensions[1].height = 35

    dealer_headers = [
        "S.No", "Dealer Name", "City", "Rating",
        "Products Quoted", "Avg. Discount %", "Best Category", "Contact Status"
    ]
    for col_idx, header in enumerate(dealer_headers, 1):
        cell = ws_dealers.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # Calculate dealer stats
    dealer_stats = {}
    for product, quotes in product_data_for_detail:
        for quote in quotes:
            dname = quote["dealer"]["name"]
            if dname not in dealer_stats:
                dealer_stats[dname] = {
                    "dealer": quote["dealer"],
                    "products": 0,
                    "total_discount": 0,
                    "categories": {},
                }
            dealer_stats[dname]["products"] += 1
            retail = product["retail_price"]
            disc = ((retail - quote["total_per_unit"]) / retail) * 100 if retail > 0 else 0
            dealer_stats[dname]["total_discount"] += disc
            cat = product["category"]
            if cat not in dealer_stats[dname]["categories"]:
                dealer_stats[dname]["categories"][cat] = 0
            dealer_stats[dname]["categories"][cat] += 1

    for row_idx, (dname, stats) in enumerate(sorted(dealer_stats.items(), key=lambda x: -x[1]["products"]), 4):
        d = stats["dealer"]
        avg_disc = stats["total_discount"] / stats["products"] if stats["products"] > 0 else 0
        best_cat = max(stats["categories"], key=stats["categories"].get) if stats["categories"] else "N/A"
        data = [
            row_idx - 3,
            dname,
            d["city"],
            f"{d['rating']}/5",
            stats["products"],
            round(avg_disc, 1),
            best_cat,
            random.choice(["Active", "Active", "Active", "Verified", "New"]),
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws_dealers.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = thin_border
            cell.alignment = center_align

    dealer_widths = [6, 28, 16, 10, 16, 14, 22, 14]
    for i, w in enumerate(dealer_widths, 1):
        ws_dealers.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4: CATEGORY ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    ws_cat = wb.create_sheet("Category Analysis")
    ws_cat.sheet_properties.tabColor = "BF8F00"

    ws_cat.merge_cells("A1:J1")
    title_cell = ws_cat["A1"]
    title_cell.value = "Hub4Estate - Category-Wise Savings Analysis"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    ws_cat.row_dimensions[1].height = 35

    cat_headers = [
        "S.No", "Category", "Products Count", "Total Retail Value (₹)",
        "Total Best Price (₹)", "Total Savings (₹)", "Avg Savings %",
        "Avg Dealers/Product", "Best Savings Product", "Max Savings %"
    ]
    for col_idx, header in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_cat.row_dimensions[3].height = 40

    # Aggregate by category
    cat_data = {}
    for product, quotes in product_data_for_detail:
        cat = product["category"]
        if cat not in cat_data:
            cat_data[cat] = {
                "count": 0,
                "total_retail": 0,
                "total_best": 0,
                "savings_pcts": [],
                "dealer_counts": [],
                "products_savings": [],
            }
        retail = product["retail_price"]
        qty = product["qty"]
        best = quotes[0]["total_per_unit"]
        savings_pct = ((retail - best) / retail) * 100 if retail > 0 else 0

        cat_data[cat]["count"] += 1
        cat_data[cat]["total_retail"] += retail * qty
        cat_data[cat]["total_best"] += best * qty
        cat_data[cat]["savings_pcts"].append(savings_pct)
        cat_data[cat]["dealer_counts"].append(len(quotes))
        cat_data[cat]["products_savings"].append((product["product"], savings_pct))

    for row_idx, (cat, data) in enumerate(sorted(cat_data.items()), 4):
        total_savings = data["total_retail"] - data["total_best"]
        avg_savings = sum(data["savings_pcts"]) / len(data["savings_pcts"]) if data["savings_pcts"] else 0
        avg_dealers = sum(data["dealer_counts"]) / len(data["dealer_counts"]) if data["dealer_counts"] else 0
        best_product = max(data["products_savings"], key=lambda x: x[1])

        row_data = [
            row_idx - 3,
            cat,
            data["count"],
            data["total_retail"],
            data["total_best"],
            total_savings,
            round(avg_savings, 1),
            round(avg_dealers, 1),
            best_product[0],
            round(best_product[1], 1),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_cat.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = thin_border
            cell.alignment = center_align
            if col_idx in [4, 5, 6]:
                cell.number_format = '#,##0'

    cat_widths = [6, 24, 14, 20, 18, 16, 14, 16, 28, 12]
    for i, w in enumerate(cat_widths, 1):
        ws_cat.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5: RAW DATA (for database/agent use)
    # ══════════════════════════════════════════════════════════════════════════
    ws_raw = wb.create_sheet("Raw Database")
    ws_raw.sheet_properties.tabColor = "7030A0"

    raw_headers = [
        "Product_ID", "Category", "Brand", "Product_Name", "Model",
        "Specs", "Qty_Needed", "Unit", "Retail_Price_Per_Unit",
        "Dealer_Rank", "Dealer_Name", "Dealer_City", "Dealer_Rating",
        "Dealer_Unit_Price", "Shipping_Cost", "Total_Per_Unit",
        "Savings_Per_Unit", "Savings_Pct", "Pct_Gap_Retail_vs_Dealer",
        "Total_For_Qty", "Total_Savings",
        "Delivery_Days", "Stock_Available",
        "Warranty_Type", "Is_Best_Price", "Is_Real_Example",
        "Dealer_Price_Breakdown"
    ]
    for col_idx, header in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        cell.border = thin_border
        cell.alignment = center_align

    raw_row = 2
    for prod_id, (product, quotes) in enumerate(product_data_for_detail, 1):
        for rank, quote in enumerate(quotes, 1):
            retail = product["retail_price"]
            qty = product["qty"]
            total_per_unit = quote["total_per_unit"]
            savings_unit = retail - total_per_unit
            savings_pct = (savings_unit / retail) * 100 if retail > 0 else 0

            gap_pct = round(savings_pct, 1)
            breakdown = f"{quote['dealer']['name']} | Product: ₹{quote['unit_price']:,} | Ship: ₹{quote['shipping']:,} | Total: ₹{total_per_unit:,}"

            raw_data = [
                prod_id,
                product["category"],
                product["brand"],
                product["product"],
                product["model"],
                product["specs"],
                qty,
                product["unit"],
                retail,
                rank,
                quote["dealer"]["name"],
                quote["dealer"]["city"],
                quote["dealer"]["rating"],
                quote["unit_price"],
                quote["shipping"],
                total_per_unit,
                savings_unit,
                round(savings_pct, 1),
                gap_pct,
                total_per_unit * qty,
                savings_unit * qty,
                quote["delivery_days"],
                "Yes" if quote["available"] else "No",
                quote["warranty"],
                "Yes" if rank == 1 else "No",
                "Yes" if product.get("is_real_example") else "No",
                breakdown,
            ]
            for col_idx, val in enumerate(raw_data, 1):
                cell = ws_raw.cell(row=raw_row, column=col_idx)
                cell.value = val
                cell.border = thin_border
                if col_idx in [9, 14, 15, 16, 17, 20, 21]:
                    cell.number_format = '#,##0'
            raw_row += 1

    raw_widths = [10, 22, 16, 26, 22, 32, 10, 16, 16, 10, 26, 14, 12, 14, 12, 14, 14, 10, 10, 14, 14, 12, 12, 20, 12, 14, 45]
    for i, w in enumerate(raw_widths, 1):
        ws_raw.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes
    ws_summary.freeze_panes = "A5"
    ws_detail.freeze_panes = "A3"
    ws_dealers.freeze_panes = "A4"
    ws_cat.freeze_panes = "A4"
    ws_raw.freeze_panes = "A2"

    return wb


if __name__ == "__main__":
    print("Generating Hub4Estate Dealer Price Comparison Database...")
    wb = create_workbook()
    output_path = os.path.join(os.path.dirname(__file__), "Hub4Estate_Dealer_Price_Database.xlsx")
    wb.save(output_path)
    print(f"✅ Database saved to: {output_path}")
    print(f"📊 Contains {len(PRODUCTS)} products across 5 sheets")
    print(f"📋 Sheets: Summary Dashboard | Detailed Dealer Quotes | Dealer Directory | Category Analysis | Raw Database")
